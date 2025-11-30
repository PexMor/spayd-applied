#!/usr/bin/env python3
"""
Generate testing example files for the Batch SPAYD Application

This script creates bilingual test data (English and Czech) with proper payment identifiers
respecting Czech banking field length constraints.

Identifier Logic & Constraints:

VS (Variable Symbol) - MAX 10 digits total:
  Format: Event prefix (6-8 digits) + Person suffix (2-4 digits)
  - Event prefix examples:
    * Date-based: '250701' = YY/MM/DD (25/07/01) - 6 digits
    * Event code: '2512' = Year+Event (2025, event 12) - 4 digits
    * Detailed: '25080' = Year+Month+Event (2025/08/event 80) - 5 digits
  - Person suffix: Usually 2 digits (01-99), max 4 digits
  - Combined examples: 
    * '25070101' = 8 digits (250701 + 01)
    * '251200001' = 9 digits (2512 + 0001)
    * '2508001' = 7 digits (25080 + 001)
  
SS (Specific Symbol) - Person identifier (7 digits):
  Format: <grp-id><yy><class-id><class-ord>
  - grp-id: 1 digit (1=students, 2=others)
  - yy: 2 digits (year entered, e.g., 23 for 2023)
  - class-id: 1-2 digits (class identifier, e.g., 01 for class 1A)
  - class-ord: Exactly 2 digits (order within class, 01-99)
  - Examples:
    * '1230101' = student(1) + entered 2023(23) + class 1(01) + position 01
    * '2240215' = other(2) + year 2024(24) + class 2(02) + position 15
  
KS (Constant Symbol) - Payment type (exactly 4 digits):
  - Can store year (e.g., '2025') or payment type
  - Standard codes: '0308' (membership), '0558' (conference), '0008' (dues)

English files (en_*):
- en_accounts.json - Bank account configurations
- en_events.json - Payment events with splits
- en_people_data_with_vs.csv - People data with pre-assigned VS
- en_people_data_without_vs.csv - People data for auto-generated VS
- en_people_data_with_vs.xlsx - Excel with VS column
- en_people_data_without_vs.xlsx - Excel without VS column

Czech files (cz_*):
- cz_accounts.json - Bankovní účty
- cz_events.json - Platební události
- cz_people_data_with_vs.csv - Osobní údaje s přiřazeným VS
- cz_people_data_without_vs.csv - Osobní údaje pro automatické VS
- cz_people_data_with_vs.xlsx - Excel s VS sloupcem
- cz_people_data_without_vs.xlsx - Excel bez VS sloupce

All CSV files are generated with UTF-8 encoding to properly support Czech characters.
The batch application has been updated to handle UTF-8 encoding correctly.
"""

import json
import csv
from datetime import datetime, timedelta
import uuid

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  Warning: openpyxl not available. Install with: uv add openpyxl")


def generate_accounts_en():
    """Generate English example bank accounts"""
    accounts = [
        {
            "id": str(uuid.uuid4()),
            "name": "Main Business Account",
            "iban": "CZ6508000000192000145399",
            "currency": "CZK",
            "logoUrl": "https://example.com/logos/business.png"
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Event Collection Account",
            "iban": "CZ9455000000001234567890",
            "currency": "CZK"
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Euro Account",
            "iban": "CZ3530300000001234567891",
            "currency": "EUR",
            "logoUrl": "https://example.com/logos/euro.png"
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Dollar Account (Testing)",
            "iban": "CZ2520100000001234567892",
            "currency": "USD"
        }
    ]
    
    with open('en_accounts.json', 'w', encoding='utf-8') as f:
        json.dump(accounts, f, indent=2, ensure_ascii=False)
    
    print("✅ Generated en_accounts.json")
    return accounts


def generate_accounts_cz():
    """Generate Czech example bank accounts"""
    accounts = [
        {
            "id": str(uuid.uuid4()),
            "name": "Hlavní firemní účet",
            "iban": "CZ6508000000192000145399",
            "currency": "CZK",
            "logoUrl": "https://example.com/logos/business.png"
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Účet pro akce",
            "iban": "CZ9455000000001234567890",
            "currency": "CZK"
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Euro účet",
            "iban": "CZ3530300000001234567891",
            "currency": "EUR",
            "logoUrl": "https://example.com/logos/euro.png"
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Dolarový účet (testovací)",
            "iban": "CZ2520100000001234567892",
            "currency": "USD"
        }
    ]
    
    with open('cz_accounts.json', 'w', encoding='utf-8') as f:
        json.dump(accounts, f, indent=2, ensure_ascii=False)
    
    print("✅ Generated cz_accounts.json")
    return accounts


def generate_events_en():
    """Generate English example events with splits"""
    today = datetime.now()
    
    events = [
        {
            "id": str(uuid.uuid4()),
            "description": "Summer Camp 2025 - Full Package",
            "vsPrefix": "250701",  # YY/MM/DD format: 25/07/01 = 6 digits
            "vsSuffixLength": 2,  # 2-digit suffix (01-99) = 8 digits total
            "ssPrefix": "",  # No SS prefix (use full SS from CSV)
            "ssSuffixLength": 0,  # Use complete SS from CSV (7 digits)
            "ksPrefix": "",  # KS loaded from CSV (year 2025)
            "ksSuffixLength": 0,  # KS is complete 4 digits from CSV
            "splits": [
                {
                    "amount": 3000,
                    "dueDate": (today + timedelta(days=30)).strftime("%Y-%m-%d")
                },
                {
                    "amount": 3000,
                    "dueDate": (today + timedelta(days=60)).strftime("%Y-%m-%d")
                },
                {
                    "amount": 2000,
                    "dueDate": (today + timedelta(days=90)).strftime("%Y-%m-%d")
                }
            ],
            "emailTemplate": """Dear {{FirstName}} {{SecondName}},

We are sending you the payment details for Summer Camp 2025.

Please find the payment schedule below, including QR codes for easy payment.

Best regards,
Organization Team"""
        },
        {
            "id": str(uuid.uuid4()),
            "description": "Conference Registration Fee",
            "vsPrefix": "2512",  # Event code: 2512 = 4 digits (Conference 2025, event 12)
            "vsSuffixLength": 4,  # 4-digit suffix = 8 digits total (25120001)
            "ssPrefix": "",  # Use full SS from CSV
            "ssSuffixLength": 0,  # Complete SS from CSV
            "ksPrefix": "0558",  # KS fixed for conferences
            "ksSuffixLength": 0,  # No suffix, KS is complete
            "splits": [
                {
                    "amount": 5000
                }
            ],
            "emailTemplate": """Dear {{FirstName}} {{SecondName}},

Thank you for registering for our conference!

Please find the payment details below, including a QR code for easy payment with your banking app.

Best regards,
Conference Team"""
        },
        {
            "id": str(uuid.uuid4()),
            "description": "Membership Dues 2025 - Annual",
            "vsPrefix": "251101",  # YY/MM/DD: 25/11/01 = 6 digits (November 2025, payment 1)
            "vsSuffixLength": 2,  # 2-digit suffix = 8 digits total (25110101)
            "ssPrefix": "",  # Use full SS from CSV
            "ssSuffixLength": 0,  # Complete SS from CSV
            "ksPrefix": "0008",  # KS fixed for membership dues
            "ksSuffixLength": 0,  # No suffix
            "splits": [
                {
                    "amount": 1200,
                    "dueDate": (today + timedelta(days=14)).strftime("%Y-%m-%d")
                }
            ],
            "emailTemplate": """Dear {{FirstName}} {{SecondName}},

This is a reminder to pay your annual membership dues for 2025.

Please find the payment details below, including a QR code for easy payment.

Thank you,
Club Team"""
        },
        {
            "id": str(uuid.uuid4()),
            "description": "Workshop - 3 Installments",
            "vsPrefix": "25080",  # Event code: 25080 = 5 digits (Workshop 2025, event 80)
            "vsSuffixLength": 3,  # 3-digit suffix = 8 digits total (25080001)
            "ssPrefix": "",  # Use full SS from CSV
            "ssSuffixLength": 0,  # Complete SS from CSV
            "ksPrefix": "",  # No default KS (loaded from CSV)
            "ksSuffixLength": 0,  # Complete KS from CSV
            "splits": [
                {
                    "amount": 1500,
                    "dueDate": (today + timedelta(days=7)).strftime("%Y-%m-%d"),
                    "vsPrefix": "250801"  # Split 1: 25/08/01 = 6 digits
                },
                {
                    "amount": 1500,
                    "dueDate": (today + timedelta(days=37)).strftime("%Y-%m-%d"),
                    "vsPrefix": "250802"  # Split 2: 25/08/02 = 6 digits
                },
                {
                    "amount": 1000,
                    "dueDate": (today + timedelta(days=67)).strftime("%Y-%m-%d"),
                    "vsPrefix": "250803",  # Split 3: 25/08/03 = 6 digits
                    "ksPrefix": "0308"  # Override with specific KS
                }
            ],
            "emailTemplate": """Hello {{FirstName}} {{SecondName}},

Thank you for registering for the workshop!

Please find the payment schedule below with QR codes for each installment.

Workshop Team"""
        },
        {
            "id": str(uuid.uuid4()),
            "description": "School Trip - Vienna 2025",
            "vsPrefix": "250615",  # YY/MM/DD: 25/06/15 = 6 digits (June 15, 2025)
            "vsSuffixLength": 2,  # 2-digit suffix = 8 digits total (25061501)
            "ssPrefix": "",  # Use full SS from CSV
            "ssSuffixLength": 0,  # Complete SS from CSV
            "ksPrefix": "0308",  # Fixed KS for school trips
            "ksSuffixLength": 0,
            "splits": [
                {
                    "amount": 2500,
                    "dueDate": (today + timedelta(days=45)).strftime("%Y-%m-%d")
                },
                {
                    "amount": 2500,
                    "dueDate": (today + timedelta(days=75)).strftime("%Y-%m-%d")
                }
            ],
            "emailTemplate": """Dear Parents of {{FirstName}} {{SecondName}},

We are sending you the payment details for the school trip to Vienna 2025.

Please find the payment schedule with QR codes for each installment below.

School"""
        }
    ]
    
    with open('en_events.json', 'w', encoding='utf-8') as f:
        json.dump(events, f, indent=2, ensure_ascii=False)
    
    print("✅ Generated en_events.json")
    return events


def generate_events_cz():
    """Generate Czech example events with splits"""
    today = datetime.now()
    
    events = [
        {
            "id": str(uuid.uuid4()),
            "description": "Letní tábor 2025 - kompletní balíček",
            "vsPrefix": "250701",  # YY/MM/DD formát: 25/07/01 = 6 číslic
            "vsSuffixLength": 2,  # 2-místný suffix (01-99) = 8 číslic celkem
            "ssPrefix": "",  # Žádný SS prefix (použít celý SS z CSV)
            "ssSuffixLength": 0,  # Použít kompletní SS z CSV (7 číslic)
            "ksPrefix": "",  # KS načten z CSV (rok 2025)
            "ksSuffixLength": 0,  # KS je kompletní 4 číslice z CSV
            "splits": [
                {
                    "amount": 3000,
                    "dueDate": (today + timedelta(days=30)).strftime("%Y-%m-%d")
                },
                {
                    "amount": 3000,
                    "dueDate": (today + timedelta(days=60)).strftime("%Y-%m-%d")
                },
                {
                    "amount": 2000,
                    "dueDate": (today + timedelta(days=90)).strftime("%Y-%m-%d")
                }
            ],
            "emailTemplate": """Vážený/á {{FirstName}} {{SecondName}},

posíláme Vám platební údaje pro letní tábor 2025.

Platební údaje naleznete níže včetně QR kódu pro snadnou platbu.

S pozdravem,
Organizační tým"""
        },
        {
            "id": str(uuid.uuid4()),
            "description": "Konferenční poplatek",
            "vsPrefix": "2512",  # Kód události: 2512 = 4 číslice (Konference 2025, akce 12)
            "vsSuffixLength": 4,  # 4-místný suffix = 8 číslic celkem (25120001)
            "ssPrefix": "",  # Použít celý SS z CSV
            "ssSuffixLength": 0,  # Kompletní SS z CSV
            "ksPrefix": "0558",  # KS fixní pro konference
            "ksSuffixLength": 0,  # Žádný suffix, KS je kompletní
            "splits": [
                {
                    "amount": 5000
                }
            ],
            "emailTemplate": """Vážený/á {{FirstName}} {{SecondName}},

děkujeme za registraci na naši konferenci!

Platební údaje naleznete níže včetně QR kódu pro snadnou platbu pomocí vaší bankovní aplikace.

S pozdravem,
Konferenční tým"""
        },
        {
            "id": str(uuid.uuid4()),
            "description": "Členský příspěvek 2025 - roční",
            "vsPrefix": "251101",  # YY/MM/DD: 25/11/01 = 6 číslic (Listopad 2025, platba 1)
            "vsSuffixLength": 2,  # 2-místný suffix = 8 číslic celkem (25110101)
            "ssPrefix": "",  # Použít celý SS z CSV
            "ssSuffixLength": 0,  # Kompletní SS z CSV
            "ksPrefix": "0008",  # KS fixní pro členské příspěvky
            "ksSuffixLength": 0,  # Žádný suffix
            "splits": [
                {
                    "amount": 1200,
                    "dueDate": (today + timedelta(days=14)).strftime("%Y-%m-%d")
                }
            ],
            "emailTemplate": """Dobrý den {{FirstName}} {{SecondName}},

zasíláme Vám výzvu k úhradě členského příspěvku na rok 2025.

Platební údaje naleznete níže včetně QR kódu pro platbu.

Děkujeme,
Klub"""
        },
        {
            "id": str(uuid.uuid4()),
            "description": "Workshop - 3 splátky",
            "vsPrefix": "25080",  # Kód události: 25080 = 5 číslic (Workshop 2025, akce 80)
            "vsSuffixLength": 3,  # 3-místný suffix = 8 číslic celkem (25080001)
            "ssPrefix": "",  # Použít celý SS z CSV
            "ssSuffixLength": 0,  # Kompletní SS z CSV
            "ksPrefix": "",  # Žádný výchozí KS (načten z CSV)
            "ksSuffixLength": 0,  # Kompletní KS z CSV
            "splits": [
                {
                    "amount": 1500,
                    "dueDate": (today + timedelta(days=7)).strftime("%Y-%m-%d"),
                    "vsPrefix": "250801"  # Splátka 1: 25/08/01 = 6 číslic
                },
                {
                    "amount": 1500,
                    "dueDate": (today + timedelta(days=37)).strftime("%Y-%m-%d"),
                    "vsPrefix": "250802"  # Splátka 2: 25/08/02 = 6 číslic
                },
                {
                    "amount": 1000,
                    "dueDate": (today + timedelta(days=67)).strftime("%Y-%m-%d"),
                    "vsPrefix": "250803",  # Splátka 3: 25/08/03 = 6 číslic
                    "ksPrefix": "0308"  # Přepsání specifickým KS
                }
            ],
            "emailTemplate": """Dobrý den {{FirstName}} {{SecondName}},

děkujeme za registraci na workshop!

Platební plán včetně QR kódů pro jednotlivé splátky naleznete níže.

Tým workshopu"""
        },
        {
            "id": str(uuid.uuid4()),
            "description": "Školní zájezd - Vídeň 2025",
            "vsPrefix": "250615",  # YY/MM/DD: 25/06/15 = 6 číslic (15. června 2025)
            "vsSuffixLength": 2,  # 2-místný suffix = 8 číslic celkem (25061501)
            "ssPrefix": "",  # Použít celý SS z CSV
            "ssSuffixLength": 0,  # Kompletní SS z CSV
            "ksPrefix": "0308",  # Fixní KS pro školní zájezdy
            "ksSuffixLength": 0,  # Žádný suffix
            "splits": [
                {
                    "amount": 2500,
                    "dueDate": (today + timedelta(days=45)).strftime("%Y-%m-%d")
                },
                {
                    "amount": 2500,
                    "dueDate": (today + timedelta(days=75)).strftime("%Y-%m-%d")
                }
            ],
            "emailTemplate": """Vážení rodiče {{FirstName}} {{SecondName}},

zasíláme Vám platební údaje pro školní zájezd do Vídně.

Platební údaje včetně QR kódů pro jednotlivé splátky naleznete níže.

Škola"""
        }
    ]
    
    with open('cz_events.json', 'w', encoding='utf-8') as f:
        json.dump(events, f, indent=2, ensure_ascii=False)
    
    print("✅ Generated cz_events.json")
    return events


def generate_people_data_with_vs_cz():
    """Generate Czech people data CSV with Variable Symbol column
    
    VS: 2-digit suffix only (01-15), will be combined with event prefix
        Max total: 10 digits (event prefix 6-8 digits + suffix 2-4 digits)
    SS: Person identifier <grp-id><yy><class-id><class-ord>
        - grp-id: 1 digit (1=students, 2=others)
        - yy: 2 digits (year entered, e.g., 23 for 2023)
        - class-id: 1-2 digits (class identifier)
        - class-ord: 2 digits (order within class)
        Example: 1230101 = student(1), entered 2023(23), class 1(01), position 01
    KS: Optional 4-digit payment type (can store year like 2025)
    """
    headers = ['FirstName', 'SecondName', 'Email', 'VS', 'SS', 'KS']
    
    people = [
        ['Jan', 'Novák', 'jan.novak@example.com', '01', '1230101', '2025'],
        ['Eva', 'Svobodová', 'eva.svobodova@email.cz', '02', '1230102', '2025'],
        ['Petr', 'Dvořák', 'petr.dvorak@gmail.com', '03', '1230103', '2025'],
        ['Marie', 'Procházková', 'marie.prochazkova@seznam.cz', '04', '1230204', '2025'],
        ['Tomáš', 'Veselý', 'tomas.vesely@company.com', '05', '1230205', '2025'],
        ['Anna', 'Malá', 'anna.mala@email.cz', '06', '1230106', '2025'],
        ['Lukáš', 'Horák', 'lukas.horak@example.com', '07', '1240107', '2025'],
        ['Kateřina', 'Nová', 'katerina.nova@gmail.com', '08', '1240208', '2025'],
        ['Jakub', 'Černý', 'jakub.cerny@email.cz', '09', '1240109', '2025'],
        ['Lenka', 'Růžová', 'lenka.ruzova@company.com', '10', '1240210', '2025'],
        ['Martin', 'Bílý', 'martin.bily@example.com', '11', '2230111', '2025'],
        ['Veronika', 'Zelená', 'veronika.zelena@email.cz', '12', '2230212', '2025'],
        ['David', 'Král', 'david.kral@gmail.com', '13', '2240113', '2025'],
        ['Barbora', 'Svobodná', 'barbora.svobodna@company.com', '14', '2240214', '2025'],
        ['Michal', 'Novotný', 'michal.novotny@example.com', '15', '2230115', '2025']
    ]
    
    # CSV version
    with open('cz_people_data_with_vs.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(people)
    
    print("✅ Generated cz_people_data_with_vs.csv")
    
    # Excel version
    if OPENPYXL_AVAILABLE:
        wb = Workbook()
        ws = wb.active
        ws.title = "People Data"
        
        # Write headers
        ws.append(headers)
        
        # Write data
        for row in people:
            ws.append(row)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        wb.save('cz_people_data_with_vs.xlsx')
        print("✅ Generated cz_people_data_with_vs.xlsx")
    
    return headers, people


def generate_people_data_with_vs_en():
    """Generate English people data CSV with Variable Symbol column
    
    VS: 2-digit suffix only (01-15), will be combined with event prefix
        Max total: 10 digits (event prefix 6-8 digits + suffix 2-4 digits)
    SS: Person identifier <grp-id><yy><class-id><class-ord>
        - grp-id: 1 digit (1=students, 2=others)
        - yy: 2 digits (year entered, e.g., 23 for 2023)
        - class-id: 1-2 digits (class identifier)
        - class-ord: 2 digits (order within class)
        Example: 1230101 = student(1), entered 2023(23), class 1(01), position 01
    KS: Optional 4-digit payment type (can store year like 2025)
    """
    headers = ['FirstName', 'SecondName', 'Email', 'VS', 'SS', 'KS']
    
    people = [
        ['John', 'Smith', 'john.smith@example.com', '01', '1230101', '2025'],
        ['Sarah', 'Johnson', 'sarah.johnson@email.com', '02', '1230102', '2025'],
        ['Robert', 'Williams', 'robert.williams@company.com', '03', '1230103', '2025'],
        ['Emily', 'Brown', 'emily.brown@example.com', '04', '1230204', '2025'],
        ['Michael', 'Jones', 'michael.jones@email.com', '05', '1230205', '2025'],
        ['Lisa', 'Davis', 'lisa.davis@company.com', '06', '1230106', '2025'],
        ['James', 'Miller', 'james.miller@example.com', '07', '1240107', '2025'],
        ['Jennifer', 'Wilson', 'jennifer.wilson@email.com', '08', '1240208', '2025'],
        ['William', 'Moore', 'william.moore@company.com', '09', '1240109', '2025'],
        ['Jessica', 'Taylor', 'jessica.taylor@example.com', '10', '1240210', '2025'],
        ['David', 'Anderson', 'david.anderson@email.com', '11', '2230111', '2025'],
        ['Mary', 'Thomas', 'mary.thomas@company.com', '12', '2230212', '2025'],
        ['Richard', 'Jackson', 'richard.jackson@example.com', '13', '2240113', '2025'],
        ['Patricia', 'White', 'patricia.white@email.com', '14', '2240214', '2025'],
        ['Charles', 'Harris', 'charles.harris@company.com', '15', '2230115', '2025']
    ]
    
    # CSV version
    with open('en_people_data_with_vs.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(people)
    
    print("✅ Generated en_people_data_with_vs.csv")
    
    # Excel version
    if OPENPYXL_AVAILABLE:
        wb = Workbook()
        ws = wb.active
        ws.title = "People Data"
        
        # Write headers
        ws.append(headers)
        
        # Write data
        for row in people:
            ws.append(row)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        wb.save('en_people_data_with_vs.xlsx')
        print("✅ Generated en_people_data_with_vs.xlsx")
    
    return headers, people


def generate_people_data_without_vs_en():
    """Generate English people data CSV without Variable Symbol column
    
    This version demonstrates auto-generation of VS:
    - Event supplies VS prefix (e.g., '250701' = 25/07/01 for YY/MM/DD/split)
    - VS suffix will be auto-generated from row order (01, 02, 03...)
    - Total VS must not exceed 10 digits
    SS: Person identifier <grp-id><yy><class-id><class-ord>
    KS: Optional 4-digit constant symbol (can store year)
    """
    headers = ['FirstName', 'SecondName', 'Email', 'SS', 'KS']
    
    people = [
        ['John', 'Smith', 'john.smith@example.com', '1230101', '2025'],
        ['Sarah', 'Johnson', 'sarah.johnson@email.com', '1230102', '2025'],
        ['Robert', 'Williams', 'robert.williams@company.com', '1230103', '2025'],
        ['Emily', 'Brown', 'emily.brown@example.com', '1230204', '2025'],
        ['Michael', 'Jones', 'michael.jones@email.com', '1230205', '2025'],
        ['Lisa', 'Davis', 'lisa.davis@company.com', '1230106', '2025'],
        ['James', 'Miller', 'james.miller@example.com', '1240107', '2025'],
        ['Jennifer', 'Wilson', 'jennifer.wilson@email.com', '1240208', '2025'],
        ['William', 'Moore', 'william.moore@company.com', '1240109', '2025'],
        ['Jessica', 'Taylor', 'jessica.taylor@example.com', '1240210', '2025'],
        ['David', 'Anderson', 'david.anderson@email.com', '2230111', '2025'],
        ['Mary', 'Thomas', 'mary.thomas@company.com', '2230212', '2025'],
        ['Richard', 'Jackson', 'richard.jackson@example.com', '2240113', '2025'],
        ['Patricia', 'White', 'patricia.white@email.com', '2240214', '2025'],
        ['Charles', 'Harris', 'charles.harris@company.com', '2230115', '2025'],
        ['Linda', 'Martin', 'linda.martin@example.com', '2230216', '2025'],
        ['Joseph', 'Thompson', 'joseph.thompson@email.com', '1230317', '2025'],
        ['Barbara', 'Garcia', 'barbara.garcia@company.com', '1230318', '2025'],
        ['Thomas', 'Martinez', 'thomas.martinez@example.com', '1240119', '2025'],
        ['Susan', 'Robinson', 'susan.robinson@email.com', '1240220', '2025']
    ]
    
    # CSV version
    with open('en_people_data_without_vs.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(people)
    
    print("✅ Generated en_people_data_without_vs.csv")
    
    # Excel version
    if OPENPYXL_AVAILABLE:
        wb = Workbook()
        ws = wb.active
        ws.title = "People Data"
        
        # Write headers with styling
        ws.append(headers)
        
        # Write data
        for row in people:
            ws.append(row)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save('en_people_data_without_vs.xlsx')
        print("✅ Generated en_people_data_without_vs.xlsx")
    
    return headers, people


def generate_people_data_without_vs_cz():
    """Generate Czech people data CSV without Variable Symbol column
    
    This version demonstrates auto-generation of VS:
    - Event supplies VS prefix (e.g., '250701' = 25/07/01 for YY/MM/DD/split)
    - VS suffix will be auto-generated from row order (01, 02, 03...)
    - Total VS must not exceed 10 digits
    SS: Person identifier <grp-id><yy><class-id><class-ord>
    KS: Optional 4-digit constant symbol (can store year)
    """
    headers = ['FirstName', 'SecondName', 'Email', 'SS', 'KS']
    
    people = [
        ['Jana', 'Novotná', 'jana.novotna@example.com', '1230101', '2025'],
        ['Petr', 'Svoboda', 'petr.svoboda@email.cz', '1230102', '2025'],
        ['Martina', 'Nováková', 'martina.novakova@seznam.cz', '1230103', '2025'],
        ['Tomáš', 'Dvořák', 'tomas.dvorak@example.com', '1230204', '2025'],
        ['Lucie', 'Černá', 'lucie.cerna@email.cz', '1230205', '2025'],
        ['Pavel', 'Procházka', 'pavel.prochazka@company.com', '1230106', '2025'],
        ['Kateřina', 'Kučerová', 'katerina.kucerova@gmail.com', '1240107', '2025'],
        ['Jakub', 'Veselý', 'jakub.vesely@email.cz', '1240208', '2025'],
        ['Hana', 'Horáková', 'hana.horakova@company.com', '1240109', '2025'],
        ['Martin', 'Němec', 'martin.nemec@example.com', '1240210', '2025'],
        ['Lenka', 'Marková', 'lenka.markova@email.cz', '2230111', '2025'],
        ['Michal', 'Pospíšil', 'michal.pospisil@seznam.cz', '2230212', '2025'],
        ['Barbora', 'Králová', 'barbora.kralova@example.com', '2240113', '2025'],
        ['Ondřej', 'Beneš', 'ondrej.benes@email.cz', '2240214', '2025'],
        ['Veronika', 'Růžičková', 'veronika.ruzickova@company.com', '2230115', '2025'],
        ['Jaroslav', 'Fiala', 'jaroslav.fiala@example.com', '2230216', '2025'],
        ['Kristýna', 'Malá', 'kristyna.mala@email.cz', '1230317', '2025'],
        ['Radek', 'Sedláček', 'radek.sedlacek@company.com', '1230318', '2025'],
        ['Simona', 'Doležalová', 'simona.dolezalova@gmail.com', '1240119', '2025'],
        ['Zdeněk', 'Kolář', 'zdenek.kolar@email.cz', '1240220', '2025']
    ]
    
    # CSV version
    with open('cz_people_data_without_vs.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(people)
    
    print("✅ Generated cz_people_data_without_vs.csv")
    
    # Excel version
    if OPENPYXL_AVAILABLE:
        wb = Workbook()
        ws = wb.active
        ws.title = "People Data"
        
        # Write headers with styling
        ws.append(headers)
        
        # Write data
        for row in people:
            ws.append(row)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save('cz_people_data_without_vs.xlsx')
        print("✅ Generated cz_people_data_without_vs.xlsx")
    
    return headers, people


def generate_readme():
    """Generate README explaining the test files"""
    readme_content = """# Batch SPAYD Application - Test Data Files

This directory contains testing example files for the Batch SPAYD Application.

All files are provided in two languages:
- **English files** - prefixed with `en_`
- **Czech files** - prefixed with `cz_`

## Generated Files

### 1. Accounts (Bank Account Configurations)

**English version:** `en_accounts.json`
**Czech version:** `cz_accounts.json`

**Contents:**
- 4 example accounts (CZK, EUR, USD)
- Valid Czech IBANs
- Optional logo URLs
- Account names in respective language

**Usage:** Import in Accounts section of batch application

### 2. Events (Payment Event Configurations)

**English version:** `en_events.json`
**Czech version:** `cz_events.json`

**Contents:**
- 5 different event types
- Single and multiple payment splits (1-3 splits per event)
- Different due dates
- Various VS prefixes, SS, and KS codes
- Email templates in respective language

**Event Examples:**
- Summer Camp / Letní tábor (3 installments)
- Conference Fee / Konferenční poplatek (single payment)
- Membership Dues / Členský příspěvek (with deadline)
- Workshop (3 custom splits with overrides)
- School Trip / Školní zájezd (2 payments)

**Usage:** Import in Events section of batch application

### 3. People Data Files

#### English People Data:
- `en_people_data_with_vs.csv` - 15 English names with pre-assigned VS
- `en_people_data_with_vs.xlsx` - Same data in Excel format
- `en_people_data_without_vs.csv` - 20 English names for auto-generated VS
- `en_people_data_without_vs.xlsx` - Same data in Excel format

#### Czech People Data:
- `cz_people_data_with_vs.csv` - 15 Czech names with pre-assigned VS
- `cz_people_data_with_vs.xlsx` - Same data in Excel format
- `cz_people_data_without_vs.csv` - 20 Czech names for auto-generated VS
- `cz_people_data_without_vs.xlsx` - Same data in Excel format

**Columns (with VS):** FirstName, SecondName, Email, VS, SS, KS
**Columns (without VS):** FirstName, SecondName, Email, SS, KS

**Important Notes:**
- **VS (Variable Symbol)**: Event prefix + Person identifier. Pre-assigned in "with_vs" files, auto-generated from row index in "without_vs" files.
- **SS (Specific Symbol)**: Long-term person identifier (like student ID or member number). Always present in all files.
- **KS (Constant Symbol)**: Payment type identifier (optional, 4 digits). Examples: 0308 (membership), 0558 (conference), 0008 (dues).

## How to Generate

Run the generation script:

```bash
# Using uv (recommended)
uv run python mk.py

# Or with regular Python (requires openpyxl)
pip install openpyxl
python mk.py
```

## Testing Scenarios

### Scenario 1: English Version - Basic Import with VS
1. Import `en_accounts.json` → Accounts section
2. Import `en_events.json` → Events section  
3. Import `en_people_data_with_vs.csv` → People Data section
4. Select event and generate batch payments

### Scenario 2: Czech Version - Basic Import with VS
1. Import `cz_accounts.json` → Accounts section
2. Import `cz_events.json` → Events section  
3. Import `cz_people_data_with_vs.csv` → People Data section
4. Select event and generate batch payments

### Scenario 3: Auto-Generated VS (English)
1. Import `en_accounts.json` and `en_events.json`
2. Import `en_people_data_without_vs.csv`
3. System will auto-generate VS codes
4. Generate batch payments

### Scenario 4: Auto-Generated VS (Czech)
1. Import `cz_accounts.json` and `cz_events.json`
2. Import `cz_people_data_without_vs.csv`
3. System will auto-generate VS codes
4. Generate batch payments

### Scenario 5: Excel Import
1. Import accounts and events (any language)
2. Import `*_people_data_with_vs.xlsx` or `*_people_data_without_vs.xlsx`
3. Verify Excel parsing works correctly with Czech/English names
4. Generate payments

### Scenario 6: Multiple Splits
1. Select "Summer Camp" or "Workshop" event (any language)
2. Import any people data
3. Generate batch - should create 3 separate payments per person
4. Verify different due dates and amounts

### Scenario 7: Mixed Currencies
1. Import accounts (includes EUR and USD accounts)
2. Create event linked to EUR account
3. Import people data (any language)
4. Generate payments in EUR

### Scenario 8: Mixed Languages
1. Import `en_accounts.json` and `cz_events.json`
2. Import `en_people_data_with_vs.csv`
3. Test cross-language compatibility

## Data Structure Details

### Account Schema
```json
{
  "id": "uuid",
  "name": "string",
  "iban": "CZ...",
  "currency": "CZK|EUR|USD",
  "logoUrl": "string (optional)"
}
```

### Event Schema
```json
{
  "id": "uuid",
  "description": "string",
  "vsPrefix": "string",
  "ss": "string (optional, 1-10 digits)",
  "ks": "string (optional, 4 digits)",
  "splits": [
    {
      "amount": number,
      "dueDate": "YYYY-MM-DD (optional)",
      "vsPrefix": "string (optional override)",
      "ss": "string (optional override)",
      "ks": "string (optional override)"
    }
  ],
  "emailTemplate": "string with variables"
}
```

### People Data Schema (CSV/Excel)

**With VS (pre-assigned):**
```
FirstName,SecondName,Email,VS,SS,KS
```

**Without VS (auto-generated):**
```
FirstName,SecondName,Email,SS,KS
```

**Column Descriptions:**
- `FirstName` - Person's first name (required)
- `SecondName` - Person's last/second name (required)
- `Email` - Person's email address (required)
- `VS` - Variable Symbol: Event prefix + Person identifier (optional, will be auto-generated if missing)
- `SS` - Specific Symbol: Long-term person identifier like student/member ID (required, 10 digits max)
- `KS` - Constant Symbol: Payment type identifier (optional, exactly 4 digits)

**Email Template Variables:**
- `{{FirstName}}` - Person's first name (from CSV column)
- `{{SecondName}}` - Person's second name (from CSV column)
- `{{Email}}` - Person's email address (from CSV column)
- Note: Payment details (amount, VS, IBAN, etc.) are automatically added by the application

## Validation Rules

### IBAN
- Must start with CZ
- Must be valid format

### Variable Symbol (VS)
- Max 10 digits
- Numeric only
- Must be unique per account

### Specific Symbol (SS)
- 1-10 digits
- Numeric only

### Constant Symbol (KS)
- Exactly 4 digits
- Numeric only

### Splits
- Minimum 1 split
- Maximum 3 splits per event
- Each split must have amount > 0

## Common Issues

### Issue: Import fails
- **Solution:** Ensure file encoding is UTF-8
- **Solution:** Check CSV delimiter is comma
- **Solution:** Verify first row contains headers

### Issue: Invalid IBAN error
- **Solution:** Use Czech IBAN format (CZ + 22 digits)
- **Solution:** Generate random IBAN in app if needed

### Issue: VS duplicates
- **Solution:** Ensure VS column has unique values
- **Solution:** Let app auto-generate if omitted

## File Modifications

Feel free to modify these files for your testing needs:

1. **Add more people:** Append rows to CSV/Excel files
2. **Change amounts:** Edit Amount column values
3. **Add custom events:** Modify events.json
4. **Test edge cases:** Add empty fields, special characters, etc.

## Regenerating Files

To regenerate all files with fresh UUIDs and dates:

```bash
uv run python mk.py
```

All files will be overwritten with new data.

---

**Generated:** """ + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + """
**Script:** mk.py
"""
    
    with open('README.md', 'w', encoding='utf-8') as f:
        f.write(readme_content)
    
    print("✅ Generated README.md")


def main():
    """Main function to generate all test files"""
    print("🚀 Generating Batch SPAYD Application test files...\n")
    
    # Generate all files
    print("📦 Generating accounts...")
    generate_accounts_en()
    generate_accounts_cz()
    
    print("\n📅 Generating events...")
    generate_events_en()
    generate_events_cz()
    
    print("\n👥 Generating people data...")
    generate_people_data_with_vs_en()
    generate_people_data_with_vs_cz()
    generate_people_data_without_vs_en()
    generate_people_data_without_vs_cz()
    
    print("\n📝 Generating documentation...")
    generate_readme()
    
    print("\n" + "="*60)
    print("✨ All test files generated successfully!")
    print("="*60)
    print("\nGenerated files:")
    print("\n  English (en_) versions:")
    print("    📄 en_accounts.json")
    print("    📄 en_events.json")
    print("    📄 en_people_data_with_vs.csv")
    print("    📄 en_people_data_without_vs.csv")
    
    print("\n  Czech (cz_) versions:")
    print("    📄 cz_accounts.json")
    print("    📄 cz_events.json")
    print("    📄 cz_people_data_with_vs.csv")
    print("    📄 cz_people_data_without_vs.csv")
    
    if OPENPYXL_AVAILABLE:
        print("\n  Excel files:")
        print("    📊 en_people_data_with_vs.xlsx")
        print("    📊 en_people_data_without_vs.xlsx")
        print("    📊 cz_people_data_with_vs.xlsx")
        print("    📊 cz_people_data_without_vs.xlsx")
    else:
        print("\n  ⚠️  Excel files not generated (openpyxl not available)")
        print("     Install with: uv add openpyxl")
    
    print("\n  📖 README.md")
    print("\n💡 Import these files into the batch application to test!")
    print("   URL: https://pexmor.github.io/spayd-applied/app/batch.html")


if __name__ == "__main__":
    main()

